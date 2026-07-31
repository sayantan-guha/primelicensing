#!/usr/bin/env python3
"""
hoichoi Content Licensing Microsite — Build Script
Reads Excel data files, generates all HTML into output/
Run: python build.py
"""

import os, json, re
from openpyxl import load_workbook
from datetime import datetime

# ── Paths ─────────────────────────────────────────────────────────────────
ROOT     = os.path.dirname(os.path.abspath(__file__))
DATA     = os.path.join(ROOT, 'data')
LOGOS    = os.path.join(ROOT, 'logos')
OUT      = os.path.join(ROOT, 'output')
os.makedirs(OUT, exist_ok=True)

SERIES_XL      = os.path.join(DATA, 'hoichoi_series_data_0513.xlsx')
ADAPT_XL       = os.path.join(DATA, 'hoichoi_ip_adaptation_data_0513.xlsx')
MOVIES_XL      = os.path.join(DATA, 'hoichoi_movies_data_0513_2.xlsx')

FONTS = '<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Manrope:wght@300;400;500;600;700&family=PT+Mono:wght@400&family=Hind+Siliguri:wght@400;500;600&display=swap" rel="stylesheet">'

# Logo paths (relative to output/ for HTML)
HOICHOI_LOGO = '../logos/hoichoi.png'
PRIME_VIDEO_LOGO = '../logos/prime-video.png'

# ── Helpers ────────────────────────────────────────────────────────────────
def has_img(url):
    return bool(url and 'cloudinary' in str(url) and url not in ('—', ''))

def slug(title):
    return re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')

def fmt_date(val):
    if hasattr(val, 'strftime'): return val.strftime('%B %Y')
    return str(val or '').strip()

def write(filename, html):
    path = os.path.join(OUT, filename)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  {filename}: {os.path.getsize(path)//1024}KB")

# ── Shared CSS tokens ──────────────────────────────────────────────────────
BASE_VARS = """
:root{
  --red:#d20820;--velvet:#6d0550;
  --grad:linear-gradient(-60deg,#d20820 0%,#6d0550 100%);
  --soot:#191919;--dg:#2a2a2a;--mg:#555;--lg:#888;
  --off:#f5f5f5;--white:#fff;--border:rgba(0,0,0,.08);
  --shadow:0 2px 16px rgba(0,0,0,.07);
  --fp:'Outfit',sans-serif;--fb:'Manrope',sans-serif;
  --fm:'PT Mono',monospace;--fh:'Hind Siliguri',sans-serif;
  --r-xs:6px;--r-md:16px;--r-full:9999px;
}
*,*::before,*::after{margin:0;padding:0;box-sizing:border-box;}
body{font-family:var(--fb);color:var(--soot);-webkit-font-smoothing:antialiased;}
"""

NAV_CSS = """
.nav{position:fixed;top:0;left:0;right:0;z-index:200;background:rgba(245,245,245,.96);
  backdrop-filter:blur(12px);border-bottom:1px solid var(--border);height:50px;
  display:flex;align-items:center;}
.nav-inner{max-width:1200px;margin:0 auto;padding:0 24px;display:flex;align-items:center;gap:12px;width:100%;}
.nav-brand{display:flex;align-items:center;flex-shrink:0;text-decoration:none;}
.nav-logo{height:20px;width:auto;display:block;}
.nav-list-wrap{position:relative;flex:1;min-width:0;}
.nav-list{display:flex;align-items:center;gap:3px;overflow-x:auto;scrollbar-width:none;width:100%;}
.nav-list::-webkit-scrollbar{display:none;}
.nav-list-wrap::after{content:'›';position:absolute;top:0;right:0;bottom:0;width:44px;
  background:linear-gradient(to right,transparent,rgba(245,245,245,.97) 60%);
  display:flex;align-items:center;justify-content:flex-end;padding-right:6px;
  font-size:16px;color:var(--mg);pointer-events:none;transition:opacity .3s;}
.nav-list-wrap.at-end::after{opacity:0;}
.nav-link{font-family:var(--fm);font-size:8.5px;letter-spacing:.08em;text-transform:uppercase;
  padding:5px 10px;border-radius:var(--r-full);white-space:nowrap;color:var(--dg);
  border:1px solid transparent;background:none;cursor:pointer;font-weight:500;transition:all .18s;}
.nav-link:hover{color:var(--soot);border-color:var(--border);background:var(--white);}
.nav-link.active{background:var(--grad);color:#fff;font-weight:600;border-color:transparent;}
"""

NAV_JS = """
const navList=document.getElementById('navList');
const navWrap=document.getElementById('navWrap');
function checkNav(){if(navWrap)navWrap.classList.toggle('at-end',navList.scrollLeft+navList.clientWidth>=navList.scrollWidth-8);}
if(navList){navList.addEventListener('scroll',checkNav);checkNav();}
window.addEventListener('resize',checkNav);
"""

STAGE_CSS = """
html,body{height:100%;}
.stage{position:fixed;top:50px;left:0;right:0;bottom:0;overflow:hidden;}
.acard{position:absolute;inset:12px 24px;opacity:0;pointer-events:none;
  transform:translateX(32px);transition:opacity .32s ease,transform .32s ease;}
.acard.active{opacity:1;pointer-events:auto;transform:translateX(0);}
.acard.exit-left{opacity:0;transform:translateX(-32px);transition:opacity .22s ease,transform .22s ease;}
.acard.exit-right{opacity:0;transform:translateX(32px);transition:opacity .22s ease,transform .22s ease;}
.card-inner{width:100%;height:100%;max-width:1044px;margin:0 auto;background:var(--white);
  border:1px solid var(--border);border-radius:var(--r-md);
  box-shadow:var(--shadow),0 6px 24px rgba(210,8,32,.08);
  overflow:hidden;position:relative;display:flex;flex-direction:column;}
.card-inner::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;background:var(--grad);z-index:5;}
.content-col{flex:1;min-height:0;overflow-y:auto;scrollbar-width:thin;
  scrollbar-color:rgba(0,0,0,.1) transparent;display:flex;flex-direction:column;}
.content-col::-webkit-scrollbar{width:4px;}
.content-col::-webkit-scrollbar-thumb{background:rgba(0,0,0,.1);border-radius:2px;}
.counter{position:fixed;bottom:20px;right:24px;z-index:100;font-family:var(--fm);
  font-size:9px;letter-spacing:.12em;color:var(--mg);background:rgba(255,255,255,.9);
  backdrop-filter:blur(8px);padding:4px 11px;border-radius:var(--r-full);border:1px solid var(--border);}
"""

STAGE_JS = """
const TOTAL=__TOTAL__;let current=0,animating=false;
function showCard(idx){
  if(idx<0||idx>=TOTAL||animating)return;
  animating=true;
  const cards=document.querySelectorAll('#stage .acard');
  const prev=current,goingBack=idx<current;
  if(cards[prev]){cards[prev].classList.remove('active');cards[prev].classList.add(goingBack?'exit-right':'exit-left');
    setTimeout(()=>{if(cards[prev])cards[prev].classList.remove('exit-left','exit-right');},280);}
  current=idx;
  if(cards[current]){cards[current].style.transform=goingBack?'translateX(-32px)':'translateX(32px)';
    cards[current].classList.add('active');
    requestAnimationFrame(()=>{requestAnimationFrame(()=>{if(cards[current])cards[current].style.transform='';});});
    const cc=cards[current].querySelector('.content-col');if(cc)cc.scrollTop=0;}
  document.querySelectorAll('#navList .nav-link').forEach((l,i)=>l.classList.toggle('active',i===current));
  const al=document.querySelector('#navList .nav-link.active');
  if(al)al.scrollIntoView({block:'nearest',inline:'center',behavior:'smooth'});
  const ctr=document.getElementById('counter');if(ctr)ctr.textContent=(current+1)+' / '+TOTAL;
  setTimeout(()=>{animating=false;},360);
}
"""

CONTENT_CSS = """
.hero{position:relative;width:100%;aspect-ratio:32/9;flex-shrink:0;overflow:hidden;background:#111;}
.hero img,.hero picture img{display:block;width:100%;height:100%;object-fit:cover;object-position:center 30%;}
.hero picture{display:block;width:100%;height:100%;}
.ka-ph{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;}
.ka-letter{font-family:var(--fp);font-size:80px;font-weight:800;color:rgba(255,255,255,.06);}
.hero-scrim{display:none;}
.hero-trailer{position:absolute;bottom:14px;left:18px;display:inline-flex;align-items:center;gap:8px;
  background:var(--grad);border:none;border-radius:var(--r-full);padding:8px 16px;cursor:pointer;
  text-decoration:none;font-family:var(--fp);font-size:11px;font-weight:700;color:#fff;
  box-shadow:0 4px 16px rgba(210,8,32,.3);transition:opacity .2s;z-index:10;}
.hero-trailer:hover{opacity:.88;}
.tplay{width:22px;height:22px;border-radius:50%;background:rgba(255,255,255,.22);
  display:flex;align-items:center;justify-content:center;flex-shrink:0;}
.title-block{padding:20px 28px 0;}
.badges{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px;}
.badge{font-family:var(--fp);font-size:10.5px;font-weight:600;letter-spacing:.01em;padding:4px 11px;
  border-radius:var(--r-full);display:inline-flex;align-items:center;gap:5px;white-space:nowrap;}
.bd{width:5px;height:5px;border-radius:50%;background:rgba(255,255,255,.45);flex-shrink:0;}
.t-main{font-family:var(--fp);font-size:clamp(24px,2.8vw,42px);font-weight:800;color:var(--soot);
  letter-spacing:-.04em;line-height:.92;margin-bottom:5px;}
.t-sub{font-size:13px;color:var(--mg);margin-bottom:12px;}
.rule{width:32px;height:2px;background:var(--grad);border-radius:2px;margin-bottom:12px;}
.logline{font-size:13px;color:var(--dg);line-height:1.62;max-width:640px;margin-bottom:14px;}
.meta-row{display:grid;grid-template-columns:100px 1fr;gap:10px;align-items:start;margin-bottom:10px;}
.mlbl{font-family:var(--fm);font-size:7.5px;letter-spacing:.15em;text-transform:uppercase;
  color:var(--red);font-weight:600;padding-top:2px;line-height:1.4;}
.mtxt{font-size:12px;font-weight:500;color:var(--soot);line-height:1.55;}
.theme-row{display:flex;gap:5px;flex-wrap:wrap;margin-top:10px;}
.theme{font-family:var(--fm);font-size:7.5px;letter-spacing:.09em;text-transform:uppercase;
  padding:3px 9px;border-radius:var(--r-xs);background:var(--off);border:1px solid rgba(0,0,0,.1);color:var(--dg);}
.section{padding:18px 28px 0;}
.sec-hdr{display:flex;align-items:center;gap:10px;margin-bottom:14px;}
.sec-rule{width:4px;height:22px;border-radius:2px;background:var(--grad);flex-shrink:0;}
.sec-title{font-family:var(--fp);font-size:16px;font-weight:700;color:var(--soot);letter-spacing:-.02em;}
.divl{height:1px;background:rgba(0,0,0,.07);margin:18px 28px 0;}
.card,.world-card{background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:14px;}
.card-lbl{font-family:var(--fm);font-size:7.5px;letter-spacing:.18em;text-transform:uppercase;
  color:var(--red);margin-bottom:7px;font-weight:600;}
.card-body{font-size:11.5px;color:var(--dg);line-height:1.68;}
.arch-list{display:flex;flex-direction:column;gap:7px;}
.arch{display:flex;gap:8px;align-items:flex-start;}
.arch-dot{width:5px;height:5px;border-radius:50%;background:var(--red);flex-shrink:0;margin-top:5px;}
.arch-name{font-size:11.5px;font-weight:600;color:var(--soot);line-height:1.4;}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:10px;}
.cast-grid-2{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;max-width:400px;}
.cast-grid-4{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;}
.cast-card{background:var(--white);border:1px solid var(--border);border-radius:var(--r-md);overflow:hidden;box-shadow:var(--shadow);}
.dir-card{position:relative;}.dir-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--grad);z-index:1;}
.cast-photo{width:100%;aspect-ratio:1/1;overflow:hidden;background:#e8e8e8;}
.cast-photo img{width:100%;height:100%;object-fit:cover;object-position:center top;display:block;}
.cast-ph{width:100%;height:100%;background:linear-gradient(145deg,var(--off),#e0e0e0);
  display:flex;flex-direction:column;align-items:center;justify-content:center;gap:5px;}
.cast-ph-init{font-family:var(--fp);font-size:30px;font-weight:800;color:rgba(0,0,0,.1);}
.cast-ph-lbl{font-family:var(--fm);font-size:7px;letter-spacing:.14em;text-transform:uppercase;color:rgba(0,0,0,.2);}
.cast-info{padding:10px 12px 12px;}
.cast-role-lbl{font-family:var(--fm);font-size:7px;letter-spacing:.14em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:2px;}
.cast-name{font-family:var(--fp);font-size:12px;font-weight:700;color:var(--soot);margin-bottom:1px;}
.cast-char{font-size:10.5px;color:var(--mg);}
.char-grid-3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:12px;}
.char-img-wrap{width:100%;aspect-ratio:16/9;border-radius:var(--r-md);overflow:hidden;background:#111;box-shadow:var(--shadow);}
.char-img-wrap img{width:100%;height:100%;object-fit:cover;object-position:center top;display:block;}
.char-img-ph{width:100%;height:100%;display:flex;align-items:center;justify-content:center;}
.char-img-ph-t{font-family:var(--fm);font-size:8px;letter-spacing:.16em;text-transform:uppercase;color:rgba(255,255,255,.2);}
.char-card{background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:14px;}
.char-role{font-family:var(--fm);font-size:7.5px;letter-spacing:.18em;text-transform:uppercase;color:var(--red);margin-bottom:4px;font-weight:600;}
.char-name{font-family:var(--fp);font-size:18px;font-weight:800;color:var(--soot);letter-spacing:-.02em;margin-bottom:2px;}
.char-actor{font-size:11px;color:var(--mg);margin-bottom:8px;}
.char-rule{width:24px;height:2px;background:var(--grad);border-radius:2px;margin-bottom:8px;}
.char-desc{font-size:11.5px;color:var(--dg);line-height:1.68;}
.char-tags{display:flex;gap:5px;flex-wrap:wrap;margin-top:8px;}
.char-tag{font-family:var(--fm);font-size:7.5px;letter-spacing:.09em;text-transform:uppercase;
  padding:3px 8px;border-radius:var(--r-xs);background:var(--white);border:1px solid rgba(0,0,0,.1);color:var(--dg);}
.moodboard{display:grid;gap:10px;}
.mood-2{grid-template-columns:1fr 1fr;}.mood-4{grid-template-columns:1fr 1fr;}
.mood-img{width:100%;aspect-ratio:16/9;border-radius:var(--r-md);overflow:hidden;background:#111;box-shadow:var(--shadow);}
.mood-img img{width:100%;height:100%;object-fit:cover;display:block;}
.bullet-card{background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:14px;}
.bullet-list{display:flex;flex-direction:column;gap:9px;}
.bullet-item{display:flex;gap:9px;align-items:flex-start;}
.bullet-dot{width:5px;height:5px;border-radius:50%;background:var(--red);flex-shrink:0;margin-top:5px;}
.bullet-text{font-size:11.5px;color:var(--dg);line-height:1.65;}
.why-card{background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:16px;}
.why-intro{font-family:var(--fp);font-size:12.5px;font-weight:600;color:var(--soot);line-height:1.5;margin-bottom:14px;}
.why-pts{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;}
.why-num{font-family:var(--fp);font-size:20px;font-weight:800;background:var(--grad);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;line-height:1;margin-bottom:4px;}
.why-t{font-family:var(--fp);font-size:11px;font-weight:700;color:var(--soot);margin-bottom:3px;}
.why-b{font-size:11px;color:var(--dg);line-height:1.58;}
.pn-wrap{padding:14px 28px 18px;}
.pn-strip{display:flex;gap:10px;}
.pn-spacer{flex:1;}
.pn-btn{display:flex;align-items:center;gap:8px;background:var(--white);border:1px solid var(--border);
  border-radius:var(--r-md);padding:8px 12px;cursor:pointer;transition:box-shadow .18s;
  text-align:left;min-width:0;flex:1;max-width:240px;}
.pn-btn:hover{box-shadow:0 4px 14px rgba(0,0,0,.08);}
.pn-prev{flex-direction:row;}.pn-next{flex-direction:row;justify-content:flex-end;margin-left:auto;}
.pn-arrow{display:flex;align-items:center;flex-shrink:0;}
.pn-arrow svg{opacity:.28;transition:opacity .18s,transform .18s;}
.pn-btn:hover .pn-arrow svg{opacity:.7;}
.pn-prev:hover .pn-arrow-l svg{transform:translateX(-2px);}
.pn-next:hover .pn-arrow svg{transform:translateX(2px);}
.pn-body{min-width:0;}.pn-body-r{text-align:right;}
.pn-eyebrow{font-family:var(--fm);font-size:7px;letter-spacing:.16em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:2px;}
.pn-title{font-family:var(--fp);font-size:12px;font-weight:800;color:var(--soot);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.pn-genre{font-family:var(--fm);font-size:7px;letter-spacing:.07em;text-transform:uppercase;color:var(--lg);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px;}
"""

MOB_RESET_CSS = """
.mobile-scroll{display:none;}
@media(max-width:680px){
  .stage,.counter{display:none !important;}
  .nav-list-wrap{display:none;}
  .mobile-scroll{display:flex;flex-direction:column;gap:20px;padding:58px 14px 48px;}
  .mob-item{background:var(--white);border-radius:var(--r-md);overflow:hidden;box-shadow:var(--shadow);}
  .mob-item .hero{aspect-ratio:16/9;height:auto;}
  .mob-item .hero img,.mob-item .hero picture img{width:100%;height:100%;object-fit:cover;object-position:center 30%;display:block;position:static;}
  .mob-item .hero picture{display:block;width:100%;height:100%;}
  .mob-item .title-block{padding:16px 16px 0;}
  .mob-item .section{padding:14px 16px 0;}
  .mob-item .divl{margin:14px 16px 0;}
  .mob-item .pn-wrap{display:none;}
  .mob-item .t-main{font-size:clamp(22px,6vw,32px);}
  .mob-item .badge{font-size:9.5px;padding:3px 9px;}
  .mob-item .why-pts,.mob-item .two-col{grid-template-columns:1fr;}
  .mob-item .char-grid-3{grid-template-columns:repeat(2,1fr);}
  .mob-item .cast-grid-4{grid-template-columns:repeat(2,1fr);}
  .mob-item .cast-grid-2{max-width:100%;}
  .mob-item .meta-row{grid-template-columns:1fr;gap:3px;}
  .mob-item .franchise-card{flex-direction:column;}
  .mob-item .fc-poster{width:100%;height:180px;}
  .mob-item .section:last-of-type{padding-bottom:18px;}
}
@media(max-width:860px) and (min-width:681px){.nav-list-wrap::after{display:none;}}
"""

FRANCHISE_CSS = """
.franchise-list{display:flex;flex-direction:column;gap:14px;}
.franchise-card{background:var(--white);border:1px solid var(--border);border-radius:var(--r-md);overflow:hidden;box-shadow:var(--shadow);display:flex;}
.fc-poster{width:120px;flex-shrink:0;background:#111;overflow:hidden;}
.fc-poster img{width:100%;height:100%;object-fit:cover;object-position:center top;display:block;}
.fc-poster-ph{width:100%;height:100%;min-height:160px;display:flex;align-items:center;justify-content:center;background:#1a1a1a;}
.fc-poster-ph span{font-family:var(--fp);font-size:32px;font-weight:800;color:rgba(255,255,255,.08);}
.fc-body{flex:1;padding:18px 22px;display:flex;flex-direction:column;justify-content:center;}
.fc-top{display:flex;align-items:center;gap:7px;margin-bottom:7px;flex-wrap:wrap;}
.fc-num{font-family:var(--fm);font-size:8px;letter-spacing:.14em;text-transform:uppercase;color:var(--lg);font-weight:600;}
.fc-badge{font-family:var(--fp);font-size:9px;font-weight:600;padding:3px 9px;border-radius:var(--r-full);}
.fc-dur{font-family:var(--fm);font-size:8px;letter-spacing:.1em;text-transform:uppercase;padding:2px 8px;border-radius:var(--r-full);background:rgba(0,0,0,.06);border:1px solid rgba(0,0,0,.1);color:var(--dg);}
.fc-year{font-family:var(--fm);font-size:8px;letter-spacing:.1em;color:var(--lg);}
.fc-title{font-family:var(--fp);font-size:15px;font-weight:800;color:var(--soot);letter-spacing:-.02em;margin-bottom:3px;line-height:1.2;}
.fc-orig{font-family:var(--fh);font-size:11px;color:var(--mg);margin-bottom:8px;}
.fc-synopsis{font-size:11.5px;color:var(--dg);line-height:1.68;}
.fc-trailer{display:inline-flex;align-items:center;gap:6px;margin-top:12px;padding:6px 14px;
  border-radius:var(--r-full);background:var(--grad);color:#fff;font-family:var(--fp);
  font-size:9px;font-weight:700;text-decoration:none;align-self:flex-start;transition:opacity .18s;}
.fc-trailer:hover{opacity:.85;}
"""

# ── Component helpers ──────────────────────────────────────────────────────
ARROW = '<svg width="18" height="18" viewBox="0 0 18 18" fill="none"><path d="M3 9H15M15 9L10 4M15 9L10 14" stroke="#191919" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg>'

def sec(label):
    return f'<div class="sec-hdr"><div class="sec-rule"></div><div class="sec-title">{label}</div></div>'

def trailer_btn(url):
    if not url: return ''
    return f'<a href="{url}" target="_blank" rel="noopener" class="hero-trailer"><div class="tplay"><svg width="10" height="12" viewBox="0 0 10 12" fill="none"><path d="M1 1L9 6L1 11V1Z" fill="white"/></svg></div>Watch Trailer</a>'

def hero_img(url16, url32, title, mobile=False):
    if mobile:
        src = url16 if has_img(url16) else url32
        if has_img(src):
            return f'<img src="{src}" alt="{title}">'
        return f'<div class="ka-ph"><div class="ka-letter">{title[0]}</div></div>'
    if has_img(url32) or has_img(url16):
        mob = f'<source media="(max-width:680px)" srcset="{url16}">' if has_img(url16) else ''
        desk = url32 if has_img(url32) else url16
        return f'<picture>{mob}<img src="{desk}" alt="{title}"></picture>'
    return f'<div class="ka-ph"><div class="ka-letter">{title[0]}</div></div>'

def cast_card_html(c):
    is_dir = c['role'].lower() == 'director'
    dc = ' dir-card' if is_dir else ''
    init = ''.join(w[0] for w in c['name'].split()[:2])
    if has_img(c.get('photo','')):
        photo = f'<img src="{c["photo"]}" alt="{c["name"]}">'
    else:
        lbl = 'Director' if is_dir else 'Actor'
        photo = f'<div class="cast-ph"><div class="cast-ph-init">{init}</div><div class="cast-ph-lbl">{lbl}</div></div>'
    return f'<div class="cast-card{dc}"><div class="cast-photo">{photo}</div><div class="cast-info"><div class="cast-role-lbl">{c["role"]}</div><div class="cast-name">{c["name"]}</div><div class="cast-char">{c.get("char","")}</div></div></div>'

def bullets_html(raw):
    items = [i.strip() for i in raw.split('|') if i.strip()]
    rows = ''.join(f'<div class="bullet-item"><div class="bullet-dot"></div><div class="bullet-text">{i}</div></div>' for i in items)
    return f'<div class="bullet-card"><div class="bullet-list">{rows}</div></div>'

def why_html(intro, pts):
    if not intro: return ''
    pts_html = ''.join(f'<div><div class="why-num">0{i+1}</div><div class="why-t">{t}</div><div class="why-b">{b}</div></div>' for i,(t,b) in enumerate(pts) if t)
    return f'<div class="why-card"><div class="why-intro">{intro}</div><div class="why-pts">{pts_html}</div></div>'

def pn_strip(items, idx, onclick='showCard'):
    prev_btn = next_btn = ''
    if idx > 0:
        p = items[idx-1]
        g = p.get('genre', p.get('genre',''))[:45]
        prev_btn = f'<button class="pn-btn pn-prev" onclick="{onclick}({idx-1})"><div class="pn-arrow pn-arrow-l"><svg width="16" height="16" viewBox="0 0 16 16" fill="none"><path d="M10 3L5 8L10 13" stroke="#191919" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg></div><div class="pn-body"><div class="pn-eyebrow">Previous</div><div class="pn-title">{p["title"]}</div><div class="pn-genre">{g}</div></div></button>'
    if idx < len(items)-1:
        n = items[idx+1]
        g = n.get('genre', n.get('genre',''))[:45]
        next_btn = f'<button class="pn-btn pn-next" onclick="{onclick}({idx+1})"><div class="pn-body pn-body-r"><div class="pn-eyebrow">Next</div><div class="pn-title">{n["title"]}</div><div class="pn-genre">{g}</div></div><div class="pn-arrow"><svg width="16" height="16" viewBox="0 0 16 16" fill="none"><path d="M6 3L11 8L6 13" stroke="#191919" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg></div></button>'
    spacer = '<div class="pn-spacer"></div>' if prev_btn and next_btn else ''
    return f'<div class="pn-wrap"><div class="pn-strip">{prev_btn}{spacer}{next_btn}</div></div>' if prev_btn or next_btn else ''

def badge_genre_bg(g):
    g=g.lower()
    if 'horror' in g or 'occult' in g or 'supernatural' in g: return '#4a0a0a'
    if 'thriller' in g or 'crime' in g or 'mystery' in g or 'detective' in g: return '#1a2a4a'
    if 'drama' in g or 'survival' in g or 'romance' in g or 'relationship' in g: return '#2a1a4a'
    if 'comedy' in g or 'adventure' in g: return '#3a2a00'
    if 'courtroom' in g or 'legal' in g: return '#1a3a2a'
    return '#2a2a4a'

def nav_html(items, home_href, logo_href=HOICHOI_LOGO, onclick='showCard'):
    links = '\n      '.join(f'<button class="nav-link" data-index="{i}" onclick="{onclick}({i})">{t["title"]}</button>' for i,t in enumerate(items))
    return f'''<nav class="nav">
  <div class="nav-inner">
    <a href="{home_href}" class="nav-brand"><img src="{logo_href}" alt="hoichoi" class="nav-logo"></a>
    <div class="nav-list-wrap" id="navWrap">
      <div class="nav-list" id="navList">{links}</div>
    </div>
  </div>
</nav>'''

def stage_and_mobile(cards_desktop, cards_mobile, total):
    stage = '\n'.join(cards_desktop)
    mob   = '\n'.join(cards_mobile)
    js = STAGE_JS.replace('__TOTAL__', str(total)) + NAV_JS
    return f'''<div class="stage" id="stage">{stage}</div>
<div class="counter" id="counter">1 / {total}</div>
<div class="mobile-scroll">{mob}</div>
<script>
{js}
showCard(0);
</script>'''

def page_wrap(title, css, nav, body):
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} — hoichoi</title>
{FONTS}
<style>
{BASE_VARS}
body{{background:var(--off);}}
{css}
</style>
</head>
<body>
{nav}
{body}
</body>
</html>'''

# ════════════════════════════════════════════════════════
# 1. SERIES CATALOG
# ════════════════════════════════════════════════════════
print("\n[1/5] Building series catalog...")

wb = load_workbook(SERIES_XL)
ws = wb['Series']
ws2 = wb['USP Themes']

themes_raw = {}
for row in range(5, 300):
    ref   = ws2.cell(row=row, column=2).value
    theme = ws2.cell(row=row, column=3).value
    if ref and theme:
        if ref not in themes_raw: themes_raw[ref] = []
        themes_raw[ref].append(theme)

series = []
for row in range(6, 50):
    title = ws.cell(row=row, column=2).value
    if not title: continue
    raw = ws.cell(row=row, column=16).value
    status = raw.strftime('%B %Y') if hasattr(raw, 'strftime') else str(raw or '').strip()
    hindi = ws.cell(row=row, column=4).value or ''
    series.append({
        'title':    title,
        'original': ws.cell(row=row, column=3).value or '',
        'hindi':    hindi,
        'ip_type':  ws.cell(row=row, column=5).value or '',
        'portrait': ws.cell(row=row, column=6).value or '',
        'landscape':ws.cell(row=row, column=7).value or '',
        'genre':    ws.cell(row=row, column=8).value or '',
        'duration': ws.cell(row=row, column=9).value or '',
        'universe': ws.cell(row=row, column=10).value or '',
        'hindi_note':ws.cell(row=row, column=11).value or 'Hindi title suggestive · subject to mutual agreement',
        'synopsis': ws.cell(row=row, column=12).value or '',
        'feels':    ws.cell(row=row, column=13).value or '',
        'why':      ws.cell(row=row, column=14).value or '',
        'usp':      ws.cell(row=row, column=15).value or '',
        'status':   status,
        'type':     ws.cell(row=row, column=18).value or '',
        'themes':   themes_raw.get(hindi, []),
    })

def series_card(s, idx, mobile=False):
    gb = badge_genre_bg(s['genre'])
    ip = s['ip_type'].lower()
    if 'franchise' in ip:   fb,ib = 'Franchise','var(--grad)'
    elif 'new' in ip:       fb,ib = 'New IP','#1a4a2a'
    else:                   fb,ib = 'Adapted IP','#2a2a4a'
    st = s['status'].lower()
    if 'available' in st:   sb = '#1a4a1a'
    elif '2026' in st:      sb = '#1a3a6b'
    else:                   sb = '#7a3a00'
    tp = s['type'].strip()
    if tp == 'TV+':            type_bg = '#1a3a6b'
    elif 'film' in tp.lower(): type_bg = '#6b1a1a'
    else:                      type_bg = '#333'

    if mobile:
        if has_img(s.get('landscape','')): img = f'<img src="{s["landscape"]}" alt="{s["title"]}">'
        elif has_img(s.get('portrait','')): img = f'<img src="{s["portrait"]}" alt="{s["title"]}">'
        else: img = f'<div class="ka-ph"><div class="ka-letter">{s["title"][0]}</div></div>'
    else:
        if has_img(s.get('portrait','')) and has_img(s.get('landscape','')):
            img = f'<picture><source media="(max-width:680px)" srcset="{s["landscape"]}"><img src="{s["portrait"]}" alt="{s["title"]}"></picture>'
        elif has_img(s.get('portrait','')): img = f'<img src="{s["portrait"]}" alt="{s["title"]}">'
        else: img = f'<div class="ka-ph"><div class="ka-letter">{s["title"][0]}</div></div>'

    hindi_block = f'<div class="t-hindi">{s["hindi"]}</div><div class="t-note">{s["hindi_note"]}</div>' if s['hindi'] else ''
    themes_html = '<div class="theme-row">' + ''.join(f'<span class="theme">{t}</span>' for t in s['themes']) + '</div>' if s['themes'] else ''
    def mrow(lbl, txt):
        if not txt: return ''
        return f'<div class="meta-row"><div class="mlbl">{lbl}</div><div class="mtxt">{txt}</div></div>'
    usp_row = f'<div class="meta-row"><div class="mlbl">What makes it special</div><div class="mtxt">{s["usp"]}</div></div>' if s['usp'] else ''

    if not mobile:
        prev_btn = next_btn = ''
        if idx > 0:
            p = series[idx-1]
            prev_btn = f'<button class="pn-btn pn-prev" onclick="showSeries({idx-1})"><div class="pn-arrow pn-arrow-l"><svg width="16" height="16" viewBox="0 0 16 16" fill="none"><path d="M10 3L5 8L10 13" stroke="#191919" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg></div><div class="pn-body"><div class="pn-eyebrow">Previous</div><div class="pn-title">{p["title"]}</div><div class="pn-genre">{p["genre"][:45]}</div></div></button>'
        if idx < len(series)-1:
            n = series[idx+1]
            next_btn = f'<button class="pn-btn pn-next" onclick="showSeries({idx+1})"><div class="pn-body pn-body-r"><div class="pn-eyebrow">Next</div><div class="pn-title">{n["title"]}</div><div class="pn-genre">{n["genre"][:45]}</div></div><div class="pn-arrow"><svg width="16" height="16" viewBox="0 0 16 16" fill="none"><path d="M6 3L11 8L6 13" stroke="#191919" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg></div></button>'
        spacer = '<div class="pn-spacer"></div>' if prev_btn and next_btn else ''
        pn = f'<div class="nav-strip-wrap"><div class="pn-strip">{prev_btn}{spacer}{next_btn}</div></div>' if prev_btn or next_btn else ''
    else:
        pn = ''

    if mobile:
        return f'''<div class="scard">
  <div class="card">
    <div class="img-col">{img}</div>
    <div class="text-col">
      <div class="badges">
        <span class="badge b-type" style="background:{type_bg};color:#fff"><span class="bd"></span>{tp}</span>
        <span class="badge" style="background:{gb};color:#fff"><span class="bd"></span>{s["genre"]}</span>
        <span class="badge" style="background:#333;color:#fff"><span class="bd"></span>{s["duration"]}</span>
        <span class="badge" style="background:{ib};color:#fff"><span class="bd"></span>{fb}</span>
        <span class="badge" style="background:#1a1a2a;color:#fff"><span class="bd"></span>{s["universe"]}</span>
        <span class="badge" style="background:{sb};color:#fff"><span class="bd"></span>{s["status"]}</span>
      </div>
      <div class="t-main">{s["title"]}</div>
      {hindi_block}
      <div class="rule"></div>
      <div class="synopsis">{s["synopsis"]}</div>
      <div class="meta-block">{mrow("Feels like",s["feels"])}{mrow("Why it travels",s["why"])}{usp_row}</div>
      {themes_html}
    </div>
  </div>
</div>'''
    else:
        return f'''<div class="scard" id="sc{idx}" data-index="{idx}">
  <div class="card">
    <div class="img-col">{img}</div>
    <div class="text-col">
      <div class="badges">
        <span class="badge b-type" style="background:{type_bg};color:#fff"><span class="bd"></span>{tp}</span>
        <span class="badge" style="background:{gb};color:#fff"><span class="bd"></span>{s["genre"]}</span>
        <span class="badge" style="background:#333;color:#fff"><span class="bd"></span>{s["duration"]}</span>
        <span class="badge" style="background:{ib};color:#fff"><span class="bd"></span>{fb}</span>
        <span class="badge" style="background:#1a1a2a;color:#fff"><span class="bd"></span>{s["universe"]}</span>
        <span class="badge" style="background:{sb};color:#fff"><span class="bd"></span>{s["status"]}</span>
      </div>
      <div class="t-main">{s["title"]}</div>
      {hindi_block}
      <div class="rule"></div>
      <div class="synopsis">{s["synopsis"]}</div>
      <div class="meta-block">{mrow("Feels like",s["feels"])}{mrow("Why it travels",s["why"])}{usp_row}</div>
      {themes_html}
      {pn}
    </div>
  </div>
</div>'''

SERIES_CSS = """
html,body{height:100%;}
.desktop-stage{position:fixed;top:50px;left:0;right:0;bottom:0;overflow:hidden;display:flex;align-items:center;justify-content:center;padding:16px 24px;}
.scard{position:absolute;inset:16px 24px;display:flex;align-items:center;justify-content:center;opacity:0;pointer-events:none;transform:translateX(40px);transition:opacity .35s ease,transform .35s ease;}
.scard.active{opacity:1;pointer-events:auto;transform:translateX(0);}
.scard.exit-left{opacity:0;transform:translateX(-40px);transition:opacity .25s ease,transform .25s ease;}
.scard.exit-right{opacity:0;transform:translateX(40px);transition:opacity .25s ease,transform .25s ease;}
.card{background:var(--white);border:1px solid var(--border);border-radius:var(--r-md);
  box-shadow:var(--shadow),0 8px 32px rgba(210,8,32,.08);overflow:hidden;position:relative;
  display:flex;flex-direction:row;width:100%;max-width:1044px;
  height:calc(100vh - 50px - 32px);max-height:700px;min-height:400px;}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;background:var(--grad);z-index:5;}
.img-col{width:40%;flex-shrink:0;position:relative;overflow:hidden;background:#111;}
.img-col picture,.img-col picture img,.img-col > img{position:absolute;top:0;left:0;width:100%;height:100%;object-fit:cover;object-position:center top;display:block;}
.text-col{flex:1;padding:24px 32px 20px 24px;display:flex;flex-direction:column;min-width:0;overflow-y:auto;scrollbar-width:thin;scrollbar-color:rgba(0,0,0,.1) transparent;}
.text-col::-webkit-scrollbar{width:4px;}
.text-col::-webkit-scrollbar-thumb{background:rgba(0,0,0,.1);border-radius:2px;}
.badges{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px;flex-shrink:0;}
.badge{font-family:var(--fp);font-size:10.5px;font-weight:600;letter-spacing:.01em;padding:4px 11px;border-radius:var(--r-full);display:inline-flex;align-items:center;gap:5px;white-space:nowrap;}
.bd{width:5px;height:5px;border-radius:50%;background:rgba(255,255,255,.45);flex-shrink:0;}
.t-main{font-family:var(--fp);font-size:clamp(20px,2.4vw,38px);font-weight:800;color:var(--soot);letter-spacing:-.04em;line-height:.92;margin-bottom:5px;flex-shrink:0;}
.t-hindi{font-family:var(--fh);font-size:14px;color:var(--mg);margin-bottom:3px;flex-shrink:0;}
.t-note{font-family:var(--fm);font-size:8px;letter-spacing:.13em;color:var(--red);text-transform:uppercase;margin-bottom:12px;font-weight:600;flex-shrink:0;}
.rule{width:32px;height:2px;background:var(--grad);border-radius:2px;margin-bottom:12px;flex-shrink:0;}
.synopsis{font-size:11.5px;color:var(--dg);line-height:1.72;margin-bottom:14px;}
.meta-block{display:flex;flex-direction:column;gap:11px;}
.meta-row{display:grid;grid-template-columns:88px 1fr;gap:10px;align-items:start;}
.mlbl{font-family:var(--fm);font-size:7.5px;letter-spacing:.15em;text-transform:uppercase;color:var(--red);font-weight:600;padding-top:2px;line-height:1.4;}
.mtxt{font-size:11.5px;font-weight:500;color:var(--soot);line-height:1.55;}
.theme-row{display:flex;gap:5px;flex-wrap:wrap;margin-top:10px;}
.theme{font-family:var(--fm);font-size:7.5px;letter-spacing:.09em;text-transform:uppercase;padding:3px 9px;border-radius:var(--r-xs);background:var(--off);border:1px solid rgba(0,0,0,.11);color:var(--dg);}
.nav-strip-wrap{margin-top:auto;padding-top:12px;flex-shrink:0;}
.pn-strip{display:flex;gap:10px;align-items:stretch;}
.pn-spacer{flex:1;}
.pn-btn{display:flex;align-items:center;gap:8px;background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:9px 13px;cursor:pointer;transition:background .18s,box-shadow .18s;text-align:left;min-width:0;flex:1;max-width:260px;}
.pn-btn:hover{background:var(--white);box-shadow:0 4px 14px rgba(0,0,0,.08);}
.pn-prev{flex-direction:row;}.pn-next{flex-direction:row;justify-content:flex-end;margin-left:auto;}
.pn-arrow{display:flex;align-items:center;flex-shrink:0;}
.pn-arrow svg{opacity:.28;transition:opacity .18s,transform .18s;}
.pn-btn:hover .pn-arrow svg{opacity:.7;}
.pn-prev:hover .pn-arrow-l svg{transform:translateX(-2px);}
.pn-next:hover .pn-arrow svg{transform:translateX(2px);}
.pn-body{min-width:0;}.pn-body-r{text-align:right;}
.pn-eyebrow{font-family:var(--fm);font-size:7.5px;letter-spacing:.16em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:2px;}
.pn-title{font-family:var(--fp);font-size:12px;font-weight:800;color:var(--soot);letter-spacing:-.01em;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.pn-genre{font-family:var(--fm);font-size:7.5px;letter-spacing:.07em;text-transform:uppercase;color:var(--lg);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px;}
.counter{position:fixed;bottom:20px;right:24px;z-index:100;font-family:var(--fm);font-size:9px;letter-spacing:.12em;color:var(--mg);background:rgba(255,255,255,.9);backdrop-filter:blur(8px);padding:4px 11px;border-radius:var(--r-full);border:1px solid var(--border);}
.mobile-scroll{display:none;}
@media(max-width:680px){
  .desktop-stage,.counter{display:none !important;}
  .nav-list-wrap{display:none;}
  .mobile-scroll{display:flex;flex-direction:column;gap:20px;padding:58px 14px 48px;}
  .mobile-scroll .scard{position:static;opacity:1;pointer-events:auto;transform:none;transition:none;display:block;padding:0;}
  .mobile-scroll .card{flex-direction:column;min-height:unset;height:auto;max-height:none;width:100%;}
  .mobile-scroll .img-col{width:100%;height:56vw;min-height:180px;max-height:300px;position:relative;flex-shrink:0;}
  .mobile-scroll .img-col picture{position:absolute;inset:0;width:100%;height:100%;}
  .mobile-scroll .img-col picture img,.mobile-scroll .img-col > img{position:absolute;top:0;left:0;width:100%;height:100%;object-fit:cover;object-position:center 20%;}
  .mobile-scroll .text-col{padding:16px 16px 18px;overflow-y:visible;}
  .mobile-scroll .t-main{font-size:clamp(20px,5.5vw,30px);}
  .mobile-scroll .badges{gap:5px;margin-bottom:10px;}
  .mobile-scroll .badge{font-size:9.5px;padding:3px 9px;}
  .mobile-scroll .meta-row{grid-template-columns:1fr;gap:3px;}
  .mobile-scroll .nav-strip-wrap{display:none;}
}
@media(max-width:860px) and (min-width:681px){.nav-list-wrap::after{display:none;}}
"""

TOTAL_S = len(series)
nav_links = '\n      '.join(f'<button class="nav-link" data-index="{i}" onclick="showSeries({i})">{s["title"]}</button>' for i,s in enumerate(series))
desktop_cards = '\n'.join(series_card(s,i,False) for i,s in enumerate(series))
mobile_cards  = '\n'.join(series_card(s,i,True)  for i,s in enumerate(series))

series_js = f"""
const TOTAL={TOTAL_S};let current=0,animating=false;
function showSeries(idx){{
  if(idx<0||idx>=TOTAL||animating)return;
  animating=true;
  const cards=document.querySelectorAll('#desktopStage .scard');
  const prev=current,goingBack=idx<current;
  if(cards[prev]){{cards[prev].classList.remove('active');cards[prev].classList.add(goingBack?'exit-right':'exit-left');setTimeout(()=>{{if(cards[prev])cards[prev].classList.remove('exit-left','exit-right');}},300);}}
  current=idx;
  if(cards[current]){{cards[current].style.transform=goingBack?'translateX(-40px)':'translateX(40px)';cards[current].classList.add('active');requestAnimationFrame(()=>{{requestAnimationFrame(()=>{{if(cards[current])cards[current].style.transform='';}})}});const tc=cards[current].querySelector('.text-col');if(tc)tc.scrollTop=0;}}
  document.querySelectorAll('#navList .nav-link').forEach((l,i)=>l.classList.toggle('active',i===current));
  const al=document.querySelector('#navList .nav-link.active');if(al)al.scrollIntoView({{block:'nearest',inline:'center',behavior:'smooth'}});
  document.getElementById('counter').textContent=(current+1)+' / '+TOTAL;
  setTimeout(()=>{{animating=false;}},380);
}}
showSeries(0);
{NAV_JS}
"""

series_html = page_wrap(
    "hoichoi Originals — Series Catalog",
    SERIES_CSS,
    f'''<nav class="nav"><div class="nav-inner">
  <a href="../index.html" class="nav-brand"><img src="{HOICHOI_LOGO}" alt="hoichoi" class="nav-logo"></a>
  <div class="nav-list-wrap" id="navWrap"><div class="nav-list" id="navList">{nav_links}</div></div>
</div></nav>''',
    f'''<div class="desktop-stage" id="desktopStage">{desktop_cards}</div>
<div class="counter" id="counter">1 / {TOTAL_S}</div>
<div class="mobile-scroll" id="mobileScroll">{mobile_cards}</div>
<script>{series_js}</script>'''
)
write('hoichoi_series_catalog_v6.html', series_html)

# ════════════════════════════════════════════════════════
# 2. ADAPTATIONS
# ════════════════════════════════════════════════════════
print("\n[2/5] Building adaptations...")

wb = load_workbook(ADAPT_XL)
ws = wb['IP Adaptation']
ws_cc = wb['Core Character']

adapt_titles = []
for row in range(6, 20):
    t = ws.cell(row=row, column=2).value
    if not t: continue
    adapt_titles.append({
        'title':      t,
        'subtitle':   ws.cell(row=row, column=4).value or '',
        'ip_type':    ws.cell(row=row, column=5).value or '',
        'img16':      ws.cell(row=row, column=6).value or '',
        'img32':      ws.cell(row=row, column=7).value or '',
        'genre':      ws.cell(row=row, column=8).value or '',
        'form':       ws.cell(row=row, column=9).value or '',
        'universe':   ws.cell(row=row, column=10).value or '',
        'logline':    ws.cell(row=row, column=11).value or '',
        'feels':      ws.cell(row=row, column=12).value or '',
        'wms':        ws.cell(row=row, column=13).value or '',
        'world':      ws.cell(row=row, column=14).value or '',
        'archetypes': ws.cell(row=row, column=15).value or '',
        'seasonality':ws.cell(row=row, column=16).value or '',
        'adapt_notes':ws.cell(row=row, column=17).value or '',
        'trailer':    ws.cell(row=row, column=18).value or '',
        'why_intro':  ws.cell(row=row, column=19).value or '',
        'why1t':      ws.cell(row=row, column=20).value or '',
        'why1b':      ws.cell(row=row, column=21).value or '',
        'why2t':      ws.cell(row=row, column=22).value or '',
        'why2b':      ws.cell(row=row, column=23).value or '',
        'why3t':      ws.cell(row=row, column=24).value or '',
        'why3b':      ws.cell(row=row, column=25).value or '',
    })

adapt_chars = {}
for row in range(5, 20):
    ref = ws_cc.cell(row=row, column=2).value
    if not ref: continue
    adapt_chars[ref] = {
        'name':    ws_cc.cell(row=row, column=3).value or '',
        'role':    ws_cc.cell(row=row, column=4).value or '',
        'actor':   ws_cc.cell(row=row, column=5).value or '',
        'portrait':ws_cc.cell(row=row, column=6).value or '',
        'scene':   ws_cc.cell(row=row, column=7).value or '',
        'desc':    ws_cc.cell(row=row, column=8).value or '',
        'tags':    ws_cc.cell(row=row, column=9).value or '',
    }

def adapt_form_bg(f):
    return '#3a2a00' if 'long' in f.lower() else '#1a2a3a'
def adapt_ip_bg(ip):
    return 'var(--grad)' if 'franchise' in ip.lower() else '#1a3a6b'

def adapt_card(t, idx, mobile=False):
    gb = badge_genre_bg(t['genre'])
    fb = adapt_form_bg(t['form'])
    ib = adapt_ip_bg(t['ip_type'])
    img = hero_img(t['img16'], t['img32'], t['title'], mobile)
    tr  = trailer_btn(t.get('trailer',''))
    archs = ''.join(f'<div class="arch"><div class="arch-dot"></div><div class="arch-name">{a.strip()}</div></div>' for a in t['archetypes'].split('|') if a.strip())
    c = adapt_chars.get(t['title'], {})
    char_html = ''
    if c:
        p_el = f'<img src="{c["portrait"]}" alt="{c["name"]}">' if has_img(c.get('portrait','')) else '<div class="char-img-ph"><div class="char-img-ph-t">Character Portrait</div></div>'
        s_el = f'<img src="{c["scene"]}" alt="Scene">' if has_img(c.get('scene','')) else '<div class="char-img-ph"><div class="char-img-ph-t">World Still</div></div>'
        tags = ''.join(f'<span class="char-tag">{tg.strip()}</span>' for tg in (c.get('tags','') or '').split('|') if tg.strip())
        char_html = f'''<div class="char-images"><div class="char-img-wrap">{p_el}</div><div class="char-img-wrap">{s_el}</div></div>
    <div class="char-card"><div class="char-role">{c.get("role","")}</div><div class="char-name">{c.get("name","")}</div><div class="char-actor">Played by {c.get("actor","")}</div><div class="char-rule"></div><div class="char-desc">{c.get("desc","")}</div>{"<div class='char-tags'>" + tags + "</div>" if tags else ""}</div>'''
    pn = pn_strip(adapt_titles, idx) if not mobile else ''
    inner = f'''
    <div class="hero">{img}<div class="hero-scrim"></div>{tr}</div>
    <div class="title-block">
      <div class="badges">
        <span class="badge" style="background:{gb};color:#fff"><span class="bd"></span>{t["genre"]}</span>
        <span class="badge" style="background:{fb};color:#fff"><span class="bd"></span>{t["form"]}</span>
        <span class="badge" style="background:{ib};color:#fff"><span class="bd"></span>{t["ip_type"]}</span>
        <span class="badge" style="background:#1a1a2a;color:#fff"><span class="bd"></span>{t["universe"]}</span>
      </div>
      <div class="t-main">{t["title"]}</div><div class="t-sub">{t["subtitle"]}</div><div class="rule"></div>
      <div class="logline">{t["logline"]}</div>
      <div class="meta-row"><div class="mlbl">Feels like</div><div class="mtxt">{t["feels"]}</div></div>
      <div class="meta-row"><div class="mlbl">What makes it special</div><div class="mtxt">{t["wms"]}</div></div>
    </div>
    <div class="divl"></div>
    <div class="section">{sec("The Format")}<div class="two-col">
      <div class="card"><div class="card-lbl">The World</div><div class="card-body">{t["world"]}</div></div>
      <div class="card"><div class="card-lbl">Character Archetypes</div><div class="arch-list">{archs}</div></div>
    </div></div>
    <div class="divl"></div><div class="section">{sec("Core Character")}{char_html}</div>
    <div class="divl"></div><div class="section">{sec("The Opportunity")}{why_html(t["why_intro"],[(t["why1t"],t["why1b"]),(t["why2t"],t["why2b"]),(t["why3t"],t["why3b"])])}</div>
    <div class="divl"></div><div class="section">{sec("Seasonality")}{bullets_html(t["seasonality"])}</div>
    <div class="divl"></div><div class="section">{sec("Adaptation Notes")}{bullets_html(t["adapt_notes"])}</div>
    {pn}'''
    if mobile:
        return f'<div class="mob-item">{inner}</div>'
    else:
        return f'<div class="acard" id="ac{idx}" data-index="{idx}"><div class="card-inner"><div class="content-col">{inner}</div></div></div>'

ADAPT_NAV_LINKS = '\n      '.join(f'<button class="nav-link" data-index="{i}" onclick="showCard({i})">{t["title"]}</button>' for i,t in enumerate(adapt_titles))
ADAPT_CARDS = '\n'.join(adapt_card(t,i,False) for i,t in enumerate(adapt_titles))
ADAPT_MOB   = '\n'.join(adapt_card(t,i,True)  for i,t in enumerate(adapt_titles))
ADAPT_TOTAL = len(adapt_titles)

adapt_html = page_wrap(
    "hoichoi — IP Remakes & Production",
    NAV_CSS + STAGE_CSS + CONTENT_CSS + MOB_RESET_CSS,
    f'''<nav class="nav"><div class="nav-inner">
  <a href="../index.html" class="nav-brand"><img src="{HOICHOI_LOGO}" alt="hoichoi" class="nav-logo"></a>
  <div class="nav-list-wrap" id="navWrap"><div class="nav-list" id="navList">{ADAPT_NAV_LINKS}</div></div>
</div></nav>''',
    f'''<div class="stage" id="stage">{ADAPT_CARDS}</div>
<div class="counter" id="counter">1 / {ADAPT_TOTAL}</div>
<div class="mobile-scroll">{ADAPT_MOB}</div>
<script>
{STAGE_JS.replace("__TOTAL__", str(ADAPT_TOTAL))}
showCard(0);
{NAV_JS}
</script>'''
)
write('hoichoi_adaptations.html', adapt_html)

# ════════════════════════════════════════════════════════
# 3. MOVIES LANDING + CATALOG
# ════════════════════════════════════════════════════════
print("\n[3/5] Building movies landing + catalog...")

wb = load_workbook(MOVIES_XL)
ws_m   = wb['Movies']
ws_w   = wb['World']
ws_c   = wb['Cast & Director']
ws_f   = wb['Franchise Films']
ws_cc  = wb['Core Character & Moodboard']

# Movies
movies = []
for row in range(6, 20):
    t = ws_m.cell(row=row, column=2).value
    if not t: continue
    raw_status  = ws_m.cell(row=row, column=15).value
    raw_release = ws_m.cell(row=row, column=16).value
    release = fmt_date(raw_release)
    status = str(raw_status or '').strip()
    display_status = release if ('releasing' in release.lower() or 'upcoming' in release.lower()) else status
    if not display_status: display_status = status
    movies.append({
        'title':    t,
        'subtitle': ws_m.cell(row=row, column=3).value or '',
        'duration': ws_m.cell(row=row, column=5).value or '',
        'ip_type':  ws_m.cell(row=row, column=6).value or '',
        'hero32':   ws_m.cell(row=row, column=7).value or '',
        'landing':  ws_m.cell(row=row, column=8).value or '',
        'genre':    ws_m.cell(row=row, column=9).value or '',
        'universe': ws_m.cell(row=row, column=10).value or '',
        'logline':  ws_m.cell(row=row, column=11).value or '',
        'wms':      ws_m.cell(row=row, column=12).value or '',
        'themes':   ws_m.cell(row=row, column=13).value or '',
        'trailer':  ws_m.cell(row=row, column=14).value or '',
        'status':   display_status,
        'why_intro':ws_m.cell(row=row, column=17).value or '',
        'why1t':    ws_m.cell(row=row, column=18).value or '',
        'why1b':    ws_m.cell(row=row, column=19).value or '',
        'why2t':    ws_m.cell(row=row, column=20).value or '',
        'why2b':    ws_m.cell(row=row, column=21).value or '',
        'why3t':    ws_m.cell(row=row, column=22).value or '',
        'why3b':    ws_m.cell(row=row, column=23).value or '',
    })

# Fix Ei Raat status
for m in movies:
    if m['title'] == 'Ei Raat Tomar Amaar' and 'january' in m['status'].lower():
        m['status'] = 'Available Now'

worlds_m = {}
for row in range(4, 20):
    ref = ws_w.cell(row=row, column=2).value
    desc = ws_w.cell(row=row, column=3).value
    if ref and desc: worlds_m[ref] = desc

cast_m = {}
for row in range(5, 30):
    ref = ws_c.cell(row=row, column=2).value
    if not ref: continue
    if ref not in cast_m: cast_m[ref] = []
    cast_m[ref].append({
        'name':  ws_c.cell(row=row, column=3).value or '',
        'role':  ws_c.cell(row=row, column=4).value or '',
        'char':  ws_c.cell(row=row, column=5).value or '',
        'photo': ws_c.cell(row=row, column=6).value or '',
        'order': int(ws_c.cell(row=row, column=7).value or 99),
    })
for ref in cast_m:
    cast_m[ref].sort(key=lambda x: x['order'])
    for c in cast_m[ref]:
        if c['name'] == 'Debaloy Bhattacharya': c['role'] = 'Director'

franchise_f = []
for row in range(5, 20):
    num = ws_f.cell(row=row, column=2).value
    if not num: continue
    raw_yr = ws_f.cell(row=row, column=9).value
    yr = raw_yr.strftime('%Y') if hasattr(raw_yr, 'strftime') else str(raw_yr or '')
    franchise_f.append({
        'num':      num,
        'orig':     ws_f.cell(row=row, column=3).value or '',
        'eng':      ws_f.cell(row=row, column=4).value or '',
        'duration': ws_f.cell(row=row, column=5).value or '',
        'synopsis': ws_f.cell(row=row, column=6).value or '',
        'trailer':  ws_f.cell(row=row, column=7).value or '',
        'poster':   ws_f.cell(row=row, column=8).value or '',
        'year':     yr,
        'badge':    ws_f.cell(row=row, column=10).value or '',
    })

cc_m = {}
for row in range(5, 20):
    ref = ws_cc.cell(row=row, column=2).value
    if not ref: continue
    cc_m[ref] = {
        'type':  ws_cc.cell(row=row, column=3).value or '',
        'name':  ws_cc.cell(row=row, column=4).value or '',
        'role':  ws_cc.cell(row=row, column=5).value or '',
        'actor': ws_cc.cell(row=row, column=6).value or '',
        'desc':  ws_cc.cell(row=row, column=7).value or '',
        'imgs':  [ws_cc.cell(row=row, column=c).value or '' for c in range(8, 17)],
    }

def movie_card(m, idx, mobile=False):
    gb = badge_genre_bg(m['genre'])
    ib = 'var(--grad)' if 'franchise' in m['ip_type'].lower() else '#1a4a2a'
    sb_s = m['status'].lower()
    sb = '#1a4a1a' if 'available' in sb_s else '#1a3a6b' if any(x in sb_s for x in ['releasing','upcoming']) else '#333'
    img = f'<img src="{m["hero32"]}" alt="{m["title"]}">' if has_img(m.get('hero32','')) else f'<div class="ka-ph"><div class="ka-letter">{m["title"][0]}</div></div>'
    tr  = trailer_btn(m.get('trailer',''))
    themes_html = '<div class="theme-row">' + ''.join(f'<span class="theme">{t.strip()}</span>' for t in m['themes'].split('|') if t.strip()) + '</div>' if m.get('themes') else ''
    world = worlds_m.get(m['title'],'')
    world_html = f'<div class="divl"></div><div class="section">{sec("The World")}<div class="world-card"><div class="card-body">{world}</div></div></div>' if world else ''
    clist = cast_m.get(m['title'],[])
    grid = 'cast-grid-2' if len(clist)<=2 else 'cast-grid-4'
    cast_html = f'<div class="divl"></div><div class="section">{sec("Cast & Director")}<div class="{grid}">{"".join(cast_card_html(c) for c in clist)}</div></div>' if clist else ''
    cc = cc_m.get(m['title'],{})
    cc_html = ''
    if cc:
        imgs = [i for i in cc.get('imgs',[]) if has_img(i)]
        if 'core character' in cc.get('type','').lower():
            slots = ''.join(f'<div class="char-img-wrap"><img src="{u}" alt="Scene {j+1}"></div>' for j,u in enumerate(imgs[:9]))
            ct = f'<div class="char-card"><div class="char-role">{cc.get("role","")}</div><div class="char-name">{cc.get("name","")}</div><div class="char-actor">Played by {cc.get("actor","")}</div><div class="char-rule"></div><div class="char-desc">{cc.get("desc","")}</div></div>'
            cc_html = f'<div class="divl"></div><div class="section">{sec("Core Character")}<div class="char-grid-3">{slots}</div>{ct}</div>'
        else:
            mood_cols = 'mood-2' if len(imgs)<=2 else 'mood-4'
            slots = ''.join(f'<div class="mood-img"><img src="{u}" alt="Moodboard {j+1}"></div>' for j,u in enumerate(imgs))
            cc_html = f'<div class="divl"></div><div class="section">{sec("Moodboard")}<div class="moodboard {mood_cols}">{slots}</div></div>'
    fc_html = ''
    if m['title'] == 'Eken Babu':
        cards = []
        for f in franchise_f:
            upcoming = any(x in f['badge'].lower() for x in ['upcoming','releasing'])
            bbg = '#7a3a00' if upcoming else '#1a4a1a'
            tb = '' if upcoming else f'<a href="{f["trailer"]}" target="_blank" rel="noopener" class="fc-trailer"><svg width="10" height="11" viewBox="0 0 10 11" fill="none"><path d="M1 1L9 5.5L1 10V1Z" fill="white"/></svg>Watch Trailer</a>'
            poster = f'<img src="{f["poster"]}" alt="{f["eng"]}">' if has_img(f.get('poster','')) else f'<div class="fc-poster-ph"><span>{f["eng"][0]}</span></div>'
            cards.append(f'<div class="franchise-card"><div class="fc-poster">{poster}</div><div class="fc-body"><div class="fc-top"><span class="fc-num">{f["num"]}</span><span class="fc-badge" style="background:{bbg};color:#fff">{f["badge"]}</span><span class="fc-dur">{f["duration"]}</span><span class="fc-year">{f["year"]}</span></div><div class="fc-title">{f["eng"]}</div><div class="fc-orig">{f["orig"]}</div><div class="fc-synopsis">{f["synopsis"]}</div>{tb}</div></div>')
        fc_html = f'<div class="divl"></div><div class="section">{sec("The Franchise")}<div class="franchise-list">{"".join(cards)}</div></div>'
    pn = pn_strip(movies, idx) if not mobile else ''
    inner = f'''
    <div class="hero">{img}<div class="hero-scrim"></div>{tr}</div>
    <div class="title-block">
      <div class="badges">
        <span class="badge" style="background:{gb};color:#fff"><span class="bd"></span>{m["genre"]}</span>
        <span class="badge" style="background:#333;color:#fff"><span class="bd"></span>{m["duration"]}</span>
        <span class="badge" style="background:{ib};color:#fff"><span class="bd"></span>{m["ip_type"]}</span>
        <span class="badge" style="background:#1a1a2a;color:#fff"><span class="bd"></span>{m["universe"]}</span>
        <span class="badge" style="background:{sb};color:#fff"><span class="bd"></span>{m["status"]}</span>
      </div>
      <div class="t-main">{m["title"]}</div><div class="t-sub">{m["subtitle"]}</div><div class="rule"></div>
      <div class="logline">{m["logline"]}</div>
      <div class="meta-row"><div class="mlbl">What makes it special</div><div class="mtxt">{m["wms"]}</div></div>
      {themes_html}
    </div>
    {fc_html}
    {world_html}
    {cast_html}
    {cc_html}
    <div class="divl"></div>
    <div class="section">{sec("The Opportunity")}{why_html(m["why_intro"],[(m["why1t"],m["why1b"]),(m["why2t"],m["why2b"]),(m["why3t"],m["why3b"])])}</div>
    {pn}'''
    if mobile:
        return f'<div class="mob-item">{inner}</div>'
    else:
        return f'<div class="acard" id="ac{idx}" data-index="{idx}"><div class="card-inner"><div class="content-col">{inner}</div></div></div>'

MOV_NAV   = '\n      '.join(f'<button class="nav-link" data-index="{i}" onclick="showCard({i})">{m["title"]}</button>' for i,m in enumerate(movies))
MOV_CARDS = '\n'.join(movie_card(m,i,False) for i,m in enumerate(movies))
MOV_MOB   = '\n'.join(movie_card(m,i,True)  for i,m in enumerate(movies))
MOV_TOTAL = len(movies)

movies_catalog_html = page_wrap(
    "hoichoi — Bengali Films",
    NAV_CSS + STAGE_CSS + CONTENT_CSS + FRANCHISE_CSS + MOB_RESET_CSS,
    f'''<nav class="nav"><div class="nav-inner">
  <a href="movies_landing.html" class="nav-brand"><img src="{HOICHOI_LOGO}" alt="hoichoi" class="nav-logo"></a>
  <div class="nav-list-wrap" id="navWrap"><div class="nav-list" id="navList">{MOV_NAV}</div></div>
</div></nav>''',
    f'''<div class="stage" id="stage">{MOV_CARDS}</div>
<div class="counter" id="counter">1 / {MOV_TOTAL}</div>
<div class="mobile-scroll">{MOV_MOB}</div>
<script>
{STAGE_JS.replace("__TOTAL__", str(MOV_TOTAL))}
const hash=window.location.hash;
if(hash){{const idx=parseInt(hash.replace('#ac',''));if(!isNaN(idx))showCard(idx);else showCard(0);}}else{{showCard(0);}}
{NAV_JS}
</script>'''
)
write('hoichoi_movies.html', movies_catalog_html)

# Movies landing
def landing_card_html(m, idx):
    first_genre = m['genre'].split('·')[0].strip()
    img_el = f'<img src="{m["landing"]}" alt="{m["title"]}">' if has_img(m.get('landing','')) else f'<div class="land-ph">{m["title"][0]}</div>'
    flag = ''
    if any(x in m['status'].lower() for x in ['releasing','upcoming']):
        flag = f'<span class="land-flag">{m["status"]}</span>'
    elif 'franchise' in m['ip_type'].lower():
        flag = '<span class="land-flag">Franchise · 4 Films</span>'
    arrow = '<svg width="14" height="14" viewBox="0 0 14 14" fill="none"><path d="M2 7H12M12 7L8 3M12 7L8 11" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg>'
    return f'''<a href="hoichoi_movies.html#ac{idx}" class="land-card">
  <div class="land-img">{flag}{img_el}</div>
  <div class="land-info">
    <div class="land-tag">{first_genre}</div>
    <div class="land-title">{m["title"]}</div>
    <div class="land-dur">{m["duration"]}</div>
    <div class="land-arrow">View Details {arrow}</div>
  </div>
</a>'''

LAND_CSS = """
body{background:var(--off);}
.page{max-width:1100px;margin:0 auto;padding:72px 24px 64px;}
.lhead{text-align:center;margin-bottom:44px;}
.leyebrow{font-family:var(--fm);font-size:9px;letter-spacing:.22em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:10px;}
.ltitle{font-family:var(--fp);font-size:clamp(30px,5vw,48px);font-weight:800;color:var(--soot);letter-spacing:-.03em;margin-bottom:8px;}
.lsub{font-size:14px;color:var(--mg);}
.lgrid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;}
.land-card{background:var(--white);border:1px solid var(--border);border-radius:var(--r-md);overflow:hidden;box-shadow:var(--shadow);text-decoration:none;display:flex;flex-direction:column;position:relative;transition:transform .22s,box-shadow .22s;}
.land-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2.5px;background:var(--grad);z-index:1;}
.land-card:hover{transform:translateY(-4px);box-shadow:0 16px 48px rgba(0,0,0,.13);}
.land-img{width:100%;aspect-ratio:4/3;overflow:hidden;background:#111;position:relative;}
.land-img img{width:100%;height:100%;object-fit:cover;object-position:center top;display:block;transition:transform .3s;}
.land-card:hover .land-img img{transform:scale(1.04);}
.land-ph{width:100%;height:100%;display:flex;align-items:center;justify-content:center;font-family:var(--fp);font-size:52px;font-weight:800;color:rgba(255,255,255,.08);}
.land-flag{position:absolute;top:12px;left:12px;background:var(--grad);color:#fff;font-family:var(--fm);font-size:8px;letter-spacing:.12em;text-transform:uppercase;padding:4px 10px;border-radius:var(--r-full);font-weight:600;z-index:2;}
.land-info{padding:18px 20px 20px;}
.land-tag{font-family:var(--fm);font-size:8px;letter-spacing:.16em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:6px;}
.land-title{font-family:var(--fp);font-size:clamp(17px,2vw,22px);font-weight:800;color:var(--soot);letter-spacing:-.02em;line-height:1.1;margin-bottom:4px;}
.land-dur{font-family:var(--fm);font-size:8px;letter-spacing:.1em;text-transform:uppercase;color:var(--mg);margin-bottom:12px;}
.land-arrow{display:inline-flex;align-items:center;gap:6px;font-family:var(--fp);font-size:11px;font-weight:700;color:var(--red);}
.land-arrow svg{transition:transform .18s;}
.land-card:hover .land-arrow svg{transform:translateX(3px);}
@media(max-width:768px){.lgrid{grid-template-columns:1fr;gap:14px;}.page{padding:60px 16px 48px;}}
"""

movies_landing_html = page_wrap(
    "hoichoi — Bengali Films",
    NAV_CSS + LAND_CSS,
    f'<nav class="nav"><div class="nav-inner"><a href="../index.html" class="nav-brand"><img src="{HOICHOI_LOGO}" alt="hoichoi" class="nav-logo"></a></div></nav>',
    f'''<div class="page">
  <div class="lhead">
    <div class="leyebrow">hoichoi Originals</div>
    <div class="ltitle">Bengali Films</div>
    <div class="lsub">Bengali cinema for a global audience — available for acquisition and language rights.</div>
  </div>
  <div class="lgrid">{"".join(landing_card_html(m,i) for i,m in enumerate(movies))}</div>
</div>'''
)
write('movies_landing.html', movies_landing_html)

# ════════════════════════════════════════════════════════
# 4. HOMEPAGE
# ════════════════════════════════════════════════════════
print("\n[4/5] Building homepage...")

HOME_CSS = """
body{background:var(--white);}
.cover{min-height:calc(100vh - 56px);margin-top:56px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:60px 32px;background:var(--white);position:relative;overflow:hidden;}
.cover::before{content:'';position:absolute;inset:0;background:radial-gradient(ellipse 80% 60% at 50% 50%,rgba(210,8,32,.03) 0%,transparent 70%);pointer-events:none;}
.cover-inner{display:flex;flex-direction:column;align-items:center;gap:0;text-align:center;position:relative;z-index:1;}
.logo-lockup{display:flex;align-items:center;justify-content:center;gap:28px;margin-bottom:32px;}
.logo-wrap{display:flex;align-items:center;justify-content:center;}
.logo-wrap img{height:40px;width:auto;display:block;}
.logo-x{font-family:var(--fp);font-size:28px;font-weight:300;color:rgba(0,0,0,.18);letter-spacing:.05em;display:flex;align-items:center;}
.cover-rule{width:48px;height:2px;background:var(--grad);border-radius:2px;margin-bottom:28px;}
.cover-label{font-family:var(--fm);font-size:9px;letter-spacing:.28em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:14px;display:flex;align-items:center;gap:8px;}
.cover-label::before,.cover-label::after{content:'';flex:1;max-width:40px;height:1px;background:rgba(210,8,32,.3);}
.cover-title{font-family:var(--fp);font-size:clamp(22px,3.5vw,38px);font-weight:700;color:var(--soot);letter-spacing:-.03em;line-height:1.1;margin-bottom:14px;}
.cover-sub{font-size:15px;color:var(--mg);line-height:1.6;max-width:520px;margin-bottom:40px;}
.scroll-cue{display:flex;flex-direction:column;align-items:center;gap:8px;color:var(--lg);font-family:var(--fm);font-size:8px;letter-spacing:.18em;text-transform:uppercase;animation:bounce 2s infinite;}
@keyframes bounce{0%,100%{transform:translateY(0)}50%{transform:translateY(5px)}}
.scroll-cue svg{opacity:.3;}
.section-header{max-width:1100px;margin:0 auto;padding:48px 32px 0;text-align:center;}
.sh-eyebrow{font-family:var(--fm);font-size:9px;letter-spacing:.22em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:12px;}
.sh-title{font-family:var(--fp);font-size:clamp(24px,3vw,34px);font-weight:800;color:var(--soot);letter-spacing:-.03em;margin-bottom:10px;}
.blocks{max-width:1100px;margin:0 auto;padding:28px 32px 0;display:flex;flex-direction:column;gap:14px;}
.block{background:var(--white);border:1px solid var(--border);border-radius:var(--r-md);overflow:hidden;box-shadow:var(--shadow);display:grid;grid-template-columns:1fr 1.8fr;color:inherit;position:relative;min-height:220px;}
.block::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--grad);z-index:2;}
.block-left{padding:26px 26px;display:flex;flex-direction:column;justify-content:space-between;border-right:1px solid var(--border);background:var(--off);}
.block-num{font-family:var(--fp);font-size:48px;font-weight:800;background:var(--grad);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;line-height:1;margin-bottom:16px;opacity:.25;}
.block-cat{font-family:var(--fm);font-size:8px;letter-spacing:.2em;text-transform:uppercase;color:var(--red);font-weight:600;margin-bottom:8px;}
.block-title{font-family:var(--fp);font-size:clamp(20px,2.2vw,26px);font-weight:800;color:var(--soot);letter-spacing:-.03em;line-height:1.1;margin-bottom:16px;}
.block-tags{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:20px;}
.block-tag{font-family:var(--fm);font-size:8px;letter-spacing:.1em;text-transform:uppercase;padding:4px 10px;border-radius:var(--r-full);background:rgba(210,8,32,.06);border:1px solid rgba(210,8,32,.15);color:var(--red);}
.block-cta{display:inline-flex;align-items:center;gap:9px;background:var(--grad);color:#fff;font-family:var(--fp);font-size:12px;font-weight:700;padding:10px 20px;border-radius:var(--r-full);text-decoration:none;box-shadow:0 4px 16px rgba(210,8,32,.22);transition:opacity .18s,transform .18s;align-self:flex-start;}
.block-cta:hover{opacity:.9;transform:translateY(-1px);}
.block-cta svg{flex-shrink:0;transition:transform .18s;}
.block-cta:hover svg{transform:translateX(2px);}
.block-right{padding:26px 28px;display:flex;flex-direction:column;justify-content:center;gap:16px;}
.block-headline{font-family:var(--fp);font-size:clamp(14px,1.8vw,18px);font-weight:600;color:var(--soot);line-height:1.4;max-width:480px;}
.block-points{display:flex;flex-direction:column;gap:11px;}
.block-point{display:flex;gap:12px;align-items:flex-start;}
.bp-dot{width:6px;height:6px;border-radius:50%;background:var(--red);flex-shrink:0;margin-top:6px;}
.bp-text{font-size:13px;color:var(--dg);line-height:1.65;}
.bp-text strong{color:var(--soot);font-weight:600;}
.block-formats{display:flex;gap:10px;flex-wrap:wrap;padding-top:4px;}
.fmt-card{background:var(--off);border:1px solid var(--border);border-radius:var(--r-md);padding:12px 16px;flex:1;min-width:120px;}
.fmt-val{font-family:var(--fp);font-size:20px;font-weight:800;color:var(--soot);letter-spacing:-.02em;margin-bottom:3px;}
.fmt-lbl{font-family:var(--fm);font-size:8px;letter-spacing:.14em;text-transform:uppercase;color:var(--lg);}
.fmt-sub{font-size:10.5px;color:var(--mg);margin-top:3px;line-height:1.4;}
.footer{max-width:1100px;margin:40px auto 0;padding:24px 32px 40px;border-top:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:16px;}
.footer-logos{display:flex;align-items:center;gap:20px;}
.footer-logo{height:18px;width:auto;display:block;}
.footer-x{font-family:var(--fp);font-size:14px;font-weight:300;color:rgba(0,0,0,.2);}
.footer-note{font-family:var(--fm);font-size:8.5px;letter-spacing:.12em;text-transform:uppercase;color:var(--lg);}
@media(max-width:860px){
  .cover{min-height:60vh;padding:48px 24px;}
  .logo-wrap img{height:32px;}
  .logo-x{font-size:22px;}
  .blocks{padding:32px 20px 0;}
  .block{grid-template-columns:1fr;}
  .block-left{border-right:none;border-bottom:1px solid var(--border);padding:28px 24px;}
  .block-right{padding:28px 24px;}
  .block-num{font-size:36px;}
  .section-header{padding:52px 24px 0;}
}
@media(max-width:480px){
  .logo-lockup{gap:18px;}
  .logo-wrap img{height:26px;}
  .blocks{padding:24px 16px 0;}
  .footer{padding:24px 18px 36px;}
  .block-formats{flex-direction:column;}
}
"""

NAV_HOME = f'''<nav class="nav" style="position:fixed;top:0;left:0;right:0;z-index:100;background:rgba(255,255,255,.96);backdrop-filter:blur(16px);border-bottom:1px solid var(--border);height:56px;display:flex;align-items:center;">
  <div style="max-width:1100px;margin:0 auto;padding:0 32px;display:flex;align-items:center;justify-content:space-between;width:100%;">
    <a href="#" style="display:flex;align-items:center;text-decoration:none;">
      <img src="{HOICHOI_LOGO}" alt="hoichoi" style="height:22px;width:auto;display:block;">
    </a>
  </div>
</nav>'''

CTA_ARROW = '<svg width="14" height="14" viewBox="0 0 14 14" fill="none"><path d="M2 7H12M12 7L8 3M12 7L8 11" stroke="white" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg>'

home_html = page_wrap(
    "hoichoi × Prime Video — Strategic Partnership",
    HOME_CSS,
    NAV_HOME,
    f'''
<div class="cover">
  <div class="cover-inner">
    <div class="logo-lockup">
      <div class="logo-wrap"><img src="{HOICHOI_LOGO}" alt="hoichoi"></div>
      <div class="logo-x">×</div>
      <div class="logo-wrap"><img src="{PRIME_VIDEO_LOGO}" alt="Prime Video"></div>
    </div>
    <div class="cover-rule"></div>
    <div class="cover-label">Strategic Partnership</div>
    <div class="cover-title">Rooted Stories.<br>Global Reach.</div>
    <div class="cover-sub">A curated proposal across three pillars — Hindi Dubbed Originals, IP Remakes & Production, and Bengali Theatrical Films.</div>
    <div class="scroll-cue">
      <svg width="20" height="20" viewBox="0 0 20 20" fill="none"><path d="M5 8L10 13L15 8" stroke="#999" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg>
      Scroll to explore
    </div>
  </div>
</div>

<div class="section-header">
  <div class="sh-eyebrow">hoichoi Originals</div>
  <div class="sh-title">Three pillars. One partnership.</div>
</div>

<div class="blocks">
  <div class="block">
    <div class="block-left">
      <div>
        <div class="block-num">01</div>
        <div class="block-cat">Content Slate</div>
        <div class="block-title">Hindi Dubbed Originals</div>
        <div class="block-tags"><span class="block-tag">Web Series</span><span class="block-tag">TV+</span><span class="block-tag">Digital Film</span></div>
      </div>
      <a href="hoichoi_series_catalog_v6.html" class="block-cta">Explore the Slate {CTA_ARROW}</a>
    </div>
    <div class="block-right">
      <div class="block-headline">30 Originals in the next 12 months — available for Hindi dubbing and multi-language rights, along with an output deal of committed original volume for the next 2–3 years.</div>
      <div class="block-formats">
        <div class="fmt-card"><div class="fmt-val">22</div><div class="fmt-lbl">Web Series</div><div class="fmt-sub">150–250 mins · 7–10 eps · Premium episodic</div></div>
        <div class="fmt-card"><div class="fmt-val">4</div><div class="fmt-lbl">TV+</div><div class="fmt-sub">360–400 mins · 24–30 eps · Weekly format</div></div>
        <div class="fmt-card"><div class="fmt-val">4</div><div class="fmt-lbl">Digital Films</div><div class="fmt-sub">90–120 mins · Direct OTT + limited theatrical</div></div>
      </div>
      <div class="block-points">
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text">This slate covers <strong>23 titled originals</strong> — each with genre, format, synopsis, production status, and tentative release. The remaining 7 titles are currently in development and will be shared once they have gained a little more shape.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text"><strong>Proposed deal structure:</strong> 2–3 year term with a yearly originals commitment. Hindi dub as anchor right, with scope to explore multi-language rights and curated library content alongside.</div></div>
      </div>
    </div>
  </div>

  <div class="block">
    <div class="block-left">
      <div>
        <div class="block-num">02</div>
        <div class="block-cat">Format Rights</div>
        <div class="block-title">IP Remakes &amp; Production</div>
        <div class="block-tags"><span class="block-tag">Hindi Fit</span><span class="block-tag">Franchise Potential</span></div>
      </div>
      <a href="hoichoi_adaptations.html" class="block-cta">Explore IPs {CTA_ARROW}</a>
    </div>
    <div class="block-right">
      <div class="block-headline">8 curated IPs — each selected on the basis of character, strong world-building, franchise potential, and cross-market adaptability.</div>
      <div class="block-points">
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text">Curated from top-performing titles on hoichoi globally — iconic IPs with strong appeal for Hindi audiences and proven franchise mechanics.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text">For each IP: format, core character, the world of the show, what makes it special, how it builds across seasons, and early adaptation thinking.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text">Titles span horror, crime, legal drama, investigative thriller, and survival — covering the full spectrum of what travels nationally.</div></div>
      </div>
    </div>
  </div>

  <div class="block">
    <div class="block-left">
      <div>
        <div class="block-num">03</div>
        <div class="block-cat">Theatrical Films</div>
        <div class="block-title">Bengali Films</div>
        <div class="block-tags"><span class="block-tag">Bengali Original</span><span class="block-tag">Language Rights Open</span></div>
      </div>
      <a href="movies_landing.html" class="block-cta">Explore Films {CTA_ARROW}</a>
    </div>
    <div class="block-right">
      <div class="block-headline">4 proven theatrical successes and 2 upcoming releases — available in Bengali with language rights open for discussion.</div>
      <div class="block-points">
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text"><strong>The Eken franchise</strong> — 9 web series seasons + 3 theatrical films, with The Eken 4 (Kerala Conspiracy) releasing October 2026. Bengal's biggest detective universe.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text"><strong>Bhugun</strong> — starring Abhishek Banerjee, releasing June 2026. Possibly the first true Horror/Slasher film from India. Strong national streaming appeal — multi-language rights highly recommended.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text"><strong>Ei Raat Tomar Amaar</strong> — starring the iconic duo Aparna Sen and Anjan Dutt. A critical and commercial success. Available now.</div></div>
        <div class="block-point"><div class="bp-dot"></div><div class="bp-text">Theatrical films available in original Bengali. Simultaneous release with hoichoi platform post-theatrical window.</div></div>
      </div>
    </div>
  </div>
</div>

<div class="footer">
  <div class="footer-logos">
    <img src="{HOICHOI_LOGO}" alt="hoichoi" class="footer-logo">
    <span class="footer-x">×</span>
    <img src="{PRIME_VIDEO_LOGO}" alt="Prime Video" class="footer-logo">
  </div>
  <span class="footer-note">Confidential · For Discussion Purposes Only</span>
</div>'''
)
write('index.html', home_html)

print("\n✅ All 5 files generated in output/")
print("   index.html")
print("   hoichoi_series_catalog_v6.html")
print("   hoichoi_adaptations.html")
print("   movies_landing.html")
print("   hoichoi_movies.html")
